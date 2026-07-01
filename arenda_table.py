from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import re
from pathlib import Path
from typing import Mapping, Sequence

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:
	import xlrd
except ImportError:  # pragma: no cover - .xls support is optional at import time
	xlrd = None


MONTH_TO_INDEX = {
	"январь": 1,
	"февраль": 2,
	"март": 3,
	"апрель": 4,
	"май": 5,
	"июнь": 6,
	"июль": 7,
	"август": 8,
	"сентябрь": 9,
	"октябрь": 10,
	"ноябрь": 11,
	"декабрь": 12,
}

PERIOD_TITLE_RE = re.compile(r"Оборот[ы]?\s+за\s+([А-Яа-яЁё]+)\s+(\d{2,4})", re.IGNORECASE)
CONTRACT_NUMBER_RE = re.compile(r"№\s*(.+?)(?:\s+от\b|$)", re.IGNORECASE)


@dataclass(frozen=True)
class LessorFileGroup:
	lessor_name: str
	files: Sequence[str | Path]


@dataclass(frozen=True)
class PeriodKey:
	month_index: int
	year: int
	month_name: str


@dataclass
class LessorSection:
	lessor_name: str
	contract_rows: list[str]
	contract_values: dict[str, dict[PeriodKey, float]]


@dataclass
class ArendaTable:
	periods: list[PeriodKey]
	sections: list[LessorSection]


@dataclass(frozen=True)
class TaxRateRule:
	from_year: int
	multiplier: float


def _normalize_text(value: object) -> str:
	if value is None:
		return ""
	text = str(value).replace("\xa0", " ").strip()
	return " ".join(text.split())


def _coerce_number(value: object) -> float | None:
	if value in (None, ""):
		return None
	if isinstance(value, (int, float)):
		return float(value)

	text = _normalize_text(value)
	if not text:
		return None

	text = text.replace(" ", "")
	if "," in text and "." in text:
		if text.rfind(",") > text.rfind("."):
			text = text.replace(".", "").replace(",", ".")
		else:
			text = text.replace(",", "")
	else:
		text = text.replace(",", ".")

	try:
		return float(text)
	except ValueError:
		return None


def _load_sheet_rows(path: str | Path) -> list[list[object]]:
	file_path = Path(path)
	suffix = file_path.suffix.lower()

	if suffix == ".xlsx":
		workbook = openpyxl.load_workbook(file_path, data_only=True)
		sheet = workbook[workbook.sheetnames[0]]
		return [list(row) for row in sheet.iter_rows(values_only=True)]

	if suffix == ".xls":
		if xlrd is None:
			raise RuntimeError("xlrd is required to read .xls files")
		workbook = xlrd.open_workbook(file_path)
		sheet = workbook.sheet_by_index(0)
		return [sheet.row_values(row_index) for row_index in range(sheet.nrows)]

	raise ValueError(f"Unsupported workbook format: {file_path.suffix}")


def _find_cell(rows: Sequence[Sequence[object]], target: str, start_row: int = 0) -> tuple[int, int] | None:
	normalized_target = _normalize_text(target)
	for row_index in range(start_row, len(rows)):
		row = rows[row_index]
		for col_index, value in enumerate(row):
			if _normalize_text(value) == normalized_target:
				return row_index, col_index
	return None


def _find_period_key(rows: Sequence[Sequence[object]], start_row: int) -> PeriodKey | None:
	search_start = max(0, start_row - 12)
	for row_index in range(start_row - 1, search_start - 1, -1):
		for value in rows[row_index]:
			match = PERIOD_TITLE_RE.search(_normalize_text(value))
			if not match:
				continue

			month_name = match.group(1).lower()
			if month_name not in MONTH_TO_INDEX:
				continue

			year = int(match.group(2))
			if year < 100:
				year += 2000

			return PeriodKey(month_index=MONTH_TO_INDEX[month_name], year=year, month_name=month_name)
	return None


def _find_next_marker_row(rows: Sequence[Sequence[object]], start_row: int, marker: str) -> int | None:
	normalized_marker = _normalize_text(marker)
	for row_index in range(start_row, len(rows)):
		if any(_normalize_text(value) == normalized_marker for value in rows[row_index]):
			return row_index
	return None


def _extract_contract_number(text: object) -> str | None:
	normalized = _normalize_text(text)
	if not normalized:
		return None

	match = CONTRACT_NUMBER_RE.search(normalized)
	if not match:
		return None

	contract_number = match.group(1).strip().rstrip(".,;")
	return contract_number or None


def _parse_group_rows(rows: Sequence[Sequence[object]], lessee_company_name: str) -> tuple[list[PeriodKey], dict[str, dict[PeriodKey, float]]]:
	normalized_lessee = _normalize_text(lessee_company_name)
	contract_values: dict[str, dict[PeriodKey, float]] = defaultdict(lambda: defaultdict(float))
	periods_by_key: dict[tuple[int, int], PeriodKey] = {}

	search_row = 0
	while True:
		start_info = _find_cell(rows, "62.01", start_row=search_row)
		if start_info is None:
			break

		block_start = start_info[0]
		block_end = _find_next_marker_row(rows, block_start + 1, "Оборот")
		if block_end is None:
			break

		period_key = _find_period_key(rows, block_start)
		if period_key is None:
			search_row = block_end + 1
			continue

		periods_by_key[(period_key.month_index, period_key.year)] = period_key

		current_company = ""
		row_index = block_start + 1
		while row_index < block_end:
			row = rows[row_index]
			row_text = _normalize_text(row[1] if len(row) > 1 else "") or _normalize_text(row[0] if row else "")

			if not row_text or row_text in {"62", "62.01", "Оборот"}:
				row_index += 1
				continue

			if row_text.startswith("Договор"):
				contract_number = _extract_contract_number(row_text)
				if contract_number and current_company == normalized_lessee and row_index + 1 < block_end:
					next_row = rows[row_index + 1]
					next_indicator = _normalize_text(next_row[2] if len(next_row) > 2 else "")
					if next_indicator == "НУ":
						credit_value = _coerce_number(next_row[4] if len(next_row) > 4 else None)
						if credit_value is not None:
							contract_values[contract_number][period_key] += credit_value
				row_index += 2
				continue

			current_company = row_text
			row_index += 1

		search_row = block_end + 1

	periods = [periods_by_key[key] for key in sorted(periods_by_key)]
	ordered_contract_values = {contract_number: dict(values_by_period) for contract_number, values_by_period in contract_values.items()}
	return periods, ordered_contract_values


def collect_arenda_table(
	lessor_groups: Sequence[LessorFileGroup] | Mapping[str, Sequence[str | Path]],
	lessee_company_name: str,
) -> ArendaTable:
	if isinstance(lessor_groups, Mapping):
		normalized_groups = [LessorFileGroup(lessor_name=name, files=files) for name, files in lessor_groups.items()]
	else:
		normalized_groups = list(lessor_groups)

	all_periods: dict[tuple[int, int], PeriodKey] = {}
	sections: list[LessorSection] = []

	for group in normalized_groups:
		group_contracts: dict[str, dict[PeriodKey, float]] = defaultdict(lambda: defaultdict(float))
		for file_path in group.files:
			rows = _load_sheet_rows(file_path)
			periods, contract_values = _parse_group_rows(rows, lessee_company_name)
			for period in periods:
				all_periods[(period.month_index, period.year)] = period
			for contract_number, values_by_period in contract_values.items():
				for period_key, value in values_by_period.items():
					group_contracts[contract_number][period_key] += value

		sections.append(
			LessorSection(
				lessor_name=_normalize_text(group.lessor_name),
				contract_rows=list(group_contracts.keys()),
				contract_values={contract_number: dict(values_by_period) for contract_number, values_by_period in group_contracts.items()},
			)
		)

	ordered_periods = [all_periods[key] for key in sorted(all_periods)]
	return ArendaTable(periods=ordered_periods, sections=sections)


def _build_period_grid(periods: Sequence[PeriodKey]) -> list[PeriodKey]:
	months = sorted({period.month_index for period in periods})
	years = sorted({period.year for period in periods})
	period_lookup = {(period.month_index, period.year): period for period in periods}
	period_grid: list[PeriodKey] = []

	for month_index in months:
		month_name = next(period.month_name for period in periods if period.month_index == month_index)
		for year in years:
			period_grid.append(
				period_lookup.get(
					(month_index, year),
					PeriodKey(month_index=month_index, year=year, month_name=month_name),
				)
			)

	return period_grid


def _tax_multiplier_for_year(year: int, tax_rules: Sequence[TaxRateRule]) -> float:
	selected_multiplier = None
	for tax_rule in sorted(tax_rules, key=lambda item: item.from_year):
		if year >= tax_rule.from_year:
			selected_multiplier = tax_rule.multiplier

	if selected_multiplier is None:
		raise ValueError("tax_rules must contain at least one rule with from_year <= the requested year")

	return float(selected_multiplier)


def _clear_working_area(sheet: Worksheet) -> None:
	for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
		for cell in row:
			if not isinstance(cell, MergedCell):
				cell.value = None


def _write_period_headers(sheet: Worksheet, periods: Sequence[PeriodKey]) -> dict[PeriodKey, int]:
	for merged_range in list(sheet.merged_cells.ranges):
		if merged_range.min_row == 3 and merged_range.max_row == 3 and merged_range.min_col >= 3:
			sheet.unmerge_cells(str(merged_range))

	for row in (3, 4, 5):
		for column in range(3, sheet.max_column + 1):
			sheet.cell(row=row, column=column).value = None

	month_groups: list[tuple[str, list[PeriodKey]]] = []
	for period in periods:
		if month_groups and month_groups[-1][0] == period.month_name:
			month_groups[-1][1].append(period)
		else:
			month_groups.append((period.month_name, [period]))

	period_to_column: dict[PeriodKey, int] = {}
	column = 3
	green_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
	header_border = Border(
		top=Side(style="thin"),
		bottom=Side(style="thin"),
		left=Side(style="thin"),
		right=Side(style="thin"),
	)
	for month_name, month_periods in month_groups:
		start_column = column
		for period in month_periods:
			period_to_column[period] = column
			sheet.cell(row=4, column=column).value = period.year
			sheet.cell(row=5, column=column).value = "Аренда"
			sheet.cell(row=4, column=column).fill = green_fill
			sheet.cell(row=5, column=column).fill = green_fill
			sheet.cell(row=4, column=column).border = header_border
			sheet.cell(row=5, column=column).border = header_border
			column += 1
		end_column = column - 1
		for header_column in range(start_column, end_column + 1):
			sheet.cell(row=3, column=header_column).fill = green_fill
			sheet.cell(row=3, column=header_column).border = header_border
		sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=end_column)
		sheet.cell(row=3, column=start_column).value = month_name
	sheet.cell(row=3, column=start_column).fill = green_fill
	sheet.cell(row=3, column=start_column).border = header_border

	return period_to_column


def _create_arenda_workbook(periods: Sequence[PeriodKey]) -> tuple[openpyxl.Workbook, Worksheet, dict[PeriodKey, int]]:
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = "Аренда"

	period_to_column = _write_period_headers(sheet, periods)
	last_column = max(period_to_column.values(), default=3)
	last_column_letter = get_column_letter(last_column)

	sheet.merge_cells(f"D1:{last_column_letter}1")
	sheet.merge_cells(f"C2:{last_column_letter}2")
	sheet.merge_cells("A2:A5")
	sheet.merge_cells("B2:B5")
	sheet["D1"] = "Арендодатели"
	sheet["A2"] = "Контрагент"
	sheet["B2"] = "Номера Договоров"
	sheet["C2"] = "По арендованному имуществу"

	for column in range(3, last_column + 1):
		sheet.cell(row=5, column=column).value = "Аренда"

	return workbook, sheet, period_to_column


def _build_year_columns(periods: Sequence[PeriodKey], start_column: int) -> tuple[dict[int, int], int]:
	years = sorted({period.year for period in periods})
	year_to_column: dict[int, int] = {}
	column = start_column
	for year in years:
		year_to_column[year] = column
		column += 1
	return year_to_column, column - 1


def _write_year_headers(sheet: Worksheet, year_to_column: Mapping[int, int], *, start_row: int = 1) -> None:
	if not year_to_column:
		return

	blue_fill = PatternFill(fill_type="solid", fgColor="9DC3E6")
	header_border = Border(
		top=Side(style="thin"),
		bottom=Side(style="thin"),
		left=Side(style="thin"),
		right=Side(style="thin"),
	)
	start_column = min(year_to_column.values())
	end_column = max(year_to_column.values())
	sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row, end_column=end_column)
	cell = sheet.cell(row=start_row, column=start_column)
	cell.value = "ИТОГО"
	cell.fill = blue_fill
	cell.border = header_border

	for column in range(start_column, end_column + 1):
		year_cell = sheet.cell(row=4, column=column)
		year_label_cell = sheet.cell(row=5, column=column)
		year_cell.fill = blue_fill
		year_label_cell.fill = blue_fill
		year_cell.border = header_border
		year_label_cell.border = header_border

	for year, column in year_to_column.items():
		sheet.cell(row=4, column=column).value = year
		sheet.cell(row=5, column=column).value = "ИТОГО"


def _apply_row_border(sheet: Worksheet, row_number: int, start_column: int, end_column: int, *, top: str | None = None, bottom: str | None = None) -> None:
	top_side = Side(style=top) if top else Side(style=None)
	bottom_side = Side(style=bottom) if bottom else Side(style=None)
	for column in range(start_column, end_column + 1):
		cell = sheet.cell(row=row_number, column=column)
		cell.border = Border(
			top=top_side,
			bottom=bottom_side,
			left=cell.border.left,
			right=cell.border.right,
		)


def _apply_row_font(sheet: Worksheet, row_number: int, start_column: int, end_column: int, *, bold: bool = False) -> None:
	for column in range(start_column, end_column + 1):
		cell = sheet.cell(row=row_number, column=column)
		cell.font = Font(
			name=cell.font.name,
			size=cell.font.sz,
			bold=bold or cell.font.bold,
			italic=cell.font.italic,
			color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
			underline=cell.font.underline,
		)


def _apply_number_format(sheet: Worksheet, row_number: int, start_column: int, end_column: int, number_format: str) -> None:
	for column in range(start_column, end_column + 1):
		cell = sheet.cell(row=row_number, column=column)
		if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
			cell.number_format = number_format


def _format_period_row(sheet: Worksheet, row_number: int, last_column: int) -> None:
	_apply_number_format(sheet, row_number, 3, last_column, '#.##0,00')


def build_arenda_workbook(
	lessor_groups: Sequence[LessorFileGroup] | Mapping[str, Sequence[str | Path]],
	lessee_company_name: str,
	tax_rules: Sequence[TaxRateRule] | None = None,
) -> openpyxl.Workbook:
	if tax_rules is None:
		tax_rules = (
			TaxRateRule(from_year=0, multiplier=0.8),
			TaxRateRule(from_year=2025, multiplier=0.75),
		)

	table = collect_arenda_table(lessor_groups, lessee_company_name)
	period_grid = _build_period_grid(table.periods)
	workbook, sheet, period_to_column = _create_arenda_workbook(period_grid)
	year_to_column, year_last_column = _build_year_columns(period_grid, max(period_to_column.values(), default=3) + 1)
	_write_year_headers(sheet, year_to_column)

	row_number = 7
	section_period_has_values: list[dict[PeriodKey, bool]] = []
	section_year_totals: list[dict[int, float | None]] = []
	section_spans: list[tuple[int, int]] = []

	for section in table.sections:
		sheet.cell(row=row_number, column=1).value = section.lessor_name
		_apply_row_font(sheet, row_number, 1, year_last_column, bold=True)
		contract_start = row_number + 1
		contract_end = row_number
		period_has_values = {period_key: False for period_key in period_to_column}
		section_year_values = {year: 0.0 for year in year_to_column}
		section_year_has_values = {year: False for year in year_to_column}

		for contract_number in section.contract_rows:
			contract_end += 1
			sheet.cell(row=contract_end, column=2).value = contract_number
			contract_period_values = section.contract_values.get(contract_number, {})
			contract_year_values = {year: 0.0 for year in year_to_column}
			contract_year_has_values = {year: False for year in year_to_column}

			for period_key, value in contract_period_values.items():
				period_column = period_to_column[period_key]
				sheet.cell(row=contract_end, column=period_column).value = value
				period_has_values[period_key] = True
				contract_year_values[period_key.year] += float(value)
				contract_year_has_values[period_key.year] = True
				section_year_values[period_key.year] += float(value)
				section_year_has_values[period_key.year] = True

			for year, year_column in year_to_column.items():
				sheet.cell(row=contract_end, column=year_column).value = contract_year_values[year] if contract_year_has_values[year] else None

			_format_period_row(sheet, contract_end, max(period_to_column.values(), default=2))
			_apply_number_format(sheet, contract_end, min(year_to_column.values(), default=0), year_last_column, '#.##0,00')

		section_spans.append((contract_start, contract_end) if section.contract_rows else (0, 0))
		total_row = contract_end + 1
		cleaned_row = contract_end + 2
		sheet.cell(row=total_row, column=2).value = "Итого"
		sheet.cell(row=cleaned_row, column=2).value = "очищенная"

		for period_key, period_column in period_to_column.items():
			if period_has_values[period_key]:
				start_letter = get_column_letter(period_column)
				sheet.cell(row=total_row, column=period_column).value = f"=SUM({start_letter}{contract_start}:{start_letter}{contract_end})"
				sheet.cell(row=cleaned_row, column=period_column).value = f"={get_column_letter(period_column)}{total_row}*{_tax_multiplier_for_year(period_key.year, tax_rules)}"
			else:
				sheet.cell(row=total_row, column=period_column).value = None
				sheet.cell(row=cleaned_row, column=period_column).value = None

		for year, year_column in year_to_column.items():
			year_total = section_year_values[year] if section_year_has_values[year] else None
			sheet.cell(row=total_row, column=year_column).value = year_total
			sheet.cell(row=cleaned_row, column=year_column).value = f"={get_column_letter(year_column)}{total_row}*{_tax_multiplier_for_year(year, tax_rules)}" if year_total is not None else None

		_apply_row_font(sheet, total_row, 1, year_last_column, bold=True)
		_apply_row_font(sheet, cleaned_row, 1, year_last_column, bold=True)
		_apply_row_border(sheet, total_row, 1, year_last_column, top="thick")
		_apply_row_border(sheet, cleaned_row, 1, year_last_column, bottom="thick")
		_apply_number_format(sheet, total_row, 3, year_last_column, '#.##0,00')
		_apply_number_format(sheet, cleaned_row, 3, year_last_column, '#.##0,00')

		section_period_has_values.append(period_has_values)
		section_year_totals.append({year: (section_year_values[year] if section_year_has_values[year] else None) for year in year_to_column})
		row_number = cleaned_row + 2

	if table.sections:
		grand_title_row = row_number
		grand_total_row = row_number + 1
		grand_cleaned_row = row_number + 2
		sheet.cell(row=grand_title_row, column=1).value = "Итого по всем арендодателям"
		sheet.cell(row=grand_total_row, column=2).value = "Итого"
		sheet.cell(row=grand_cleaned_row, column=2).value = "очищенная"
		_apply_row_font(sheet, grand_title_row, 1, year_last_column, bold=True)
		_apply_row_font(sheet, grand_total_row, 1, year_last_column, bold=True)
		_apply_row_font(sheet, grand_cleaned_row, 1, year_last_column, bold=True)

		for period_key, period_column in period_to_column.items():
			period_letter = get_column_letter(period_column)
			if any(period_has_values.get(period_key) for period_has_values in section_period_has_values):
				ranges = [f"{period_letter}{start}:{period_letter}{end}" for start, end in section_spans if start and end]
				sheet.cell(row=grand_total_row, column=period_column).value = f"=SUM({','.join(ranges)})"
				sheet.cell(row=grand_cleaned_row, column=period_column).value = f"={period_letter}{grand_total_row}*{_tax_multiplier_for_year(period_key.year, tax_rules)}"
			else:
				sheet.cell(row=grand_total_row, column=period_column).value = None
				sheet.cell(row=grand_cleaned_row, column=period_column).value = None

		for year, year_column in year_to_column.items():
			grand_year_values = [section_year_total[year] for section_year_total in section_year_totals if section_year_total.get(year) is not None]
			grand_year_total = float(sum(grand_year_values)) if grand_year_values else None
			sheet.cell(row=grand_total_row, column=year_column).value = grand_year_total
			sheet.cell(row=grand_cleaned_row, column=year_column).value = f"={get_column_letter(year_column)}{grand_total_row}*{_tax_multiplier_for_year(year, tax_rules)}" if grand_year_total is not None else None

		_apply_row_border(sheet, grand_total_row, 1, year_last_column, top="thick")
		_apply_row_border(sheet, grand_cleaned_row, 1, year_last_column, bottom="thick")
		_apply_number_format(sheet, grand_total_row, 3, year_last_column, '#.##0,00')
		_apply_number_format(sheet, grand_cleaned_row, 3, year_last_column, '#.##0,00')

	return workbook


def save_arenda_workbook(workbook: openpyxl.Workbook, output_path: str | Path) -> None:
	workbook.save(output_path)


if __name__ == "__main__":
	workspace_root = Path.cwd()

	# Edit these settings before running the script.
	lessee_company_name = "НСС ООО"
	# Add more rules later by appending more TaxRateRule entries.
	tax_rules = (
		TaxRateRule(from_year=0, multiplier=0.8),
		TaxRateRule(from_year=2025, multiplier=0.75),
	)
	output_path = workspace_root / "Амортизация НСС 2026_аренда.xlsx"

	lessor_groups = {
		"ООО ПЕРСПЕКТИВА": [
			workspace_root / "example_table" / "Анализ счёта 90.01 за 2024 г. ООО  ПЕРСПЕКТИВА.xlsx",
			workspace_root / "example_table" / "Анализ счёта 90.01 за 2025 г. ООО  ПЕРСПЕКТИВА.xlsx",
			workspace_root / "example_table" / "Анализ счёта 90.01 за Январь 2026 г. - май 2026 г. ООО   ПЕРСПЕКТИВА.xlsx",
		],
		"ООО КОНТЕНТ": [
			workspace_root / "example_table" / "Анализ счёта 90.01 за 2024 г. ООО  КОНТЕНТ.xls",
			workspace_root / "example_table" / "Анализ счёта 90.01 за 2025 г. ООО  КОНТЕНТ.xls",
			workspace_root / "example_table" / "Анализ счёта 90.01 за 2026 г. ООО  КОНТЕНТ.xls"
		]
	}

	workbook = build_arenda_workbook(
		lessor_groups=lessor_groups,
		lessee_company_name=lessee_company_name,
		tax_rules=tax_rules,
	)
	save_arenda_workbook(workbook, output_path)
	print(f"Saved to {output_path}")
