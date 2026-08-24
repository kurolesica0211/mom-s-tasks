from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tax_rules import DEFAULT_TAX_RULES, TaxRateRule, tax_multiplier_for_year

try:
    import xlrd
except ImportError:  # pragma: no cover - .xls support is optional at import time
    xlrd = None


MONTH_NAMES = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]
MONTH_TO_INDEX = {name: index for index, name in enumerate(MONTH_NAMES, start=1)}

COUNTERPARTY_LABEL = "Контрагент"
CONTRACTS_LABEL_DEFAULT = "Договоры"
PROPERTY_LABEL = "По арендованному имуществу"
AMORTIZATION_LABEL = "Амортизация"
TOTAL_LABEL = "Итого"
CLEANED_LABEL = "очищенная"

CURRENCY_FORMAT = '_-* #,##0.00\\ _₽_-;\\-* #,##0.00\\ _₽_-;_-* "-"??\\ _₽_-;_-@_-'

_INVALID_SHEET_CHARS = set(":\\/?*[]")


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\xa0", " ").strip()
    return " ".join(text.split())


def _coerce_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = _normalize_text(value)
    if not text:
        return None

    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


@dataclass(frozen=True)
class Period:
    year: int
    month_index: int

    @property
    def month_name(self) -> str:
        return MONTH_NAMES[self.month_index - 1]


@dataclass
class ContractRow:
    contract_number: str
    values: dict[Period, object] = field(default_factory=dict)


@dataclass
class DepreciationTable:
    table_title: str
    contracts_label: str = CONTRACTS_LABEL_DEFAULT
    periods: set[Period] = field(default_factory=set)
    contracts: list[ContractRow] = field(default_factory=list)

    def find_contract(self, contract_number: str) -> ContractRow | None:
        normalized = _normalize_text(contract_number)
        for row in self.contracts:
            if _normalize_text(row.contract_number) == normalized:
                return row
        return None


@dataclass(frozen=True)
class ContractInput:
    contract_number: str
    contract_file: str | Path


@dataclass(frozen=True)
class ApplyResult:
    contract_number: str
    period: Period
    period_created: bool
    contract_created: bool
    value_written: bool


@dataclass(frozen=True)
class GenerationResult:
    workbook: openpyxl.Workbook
    sheet_name: str
    messages: list[str]
    contract_errors: list[str]
    asset_failures: list[AssetFailure]


def _safe_sheet_title(name: str) -> str:
    cleaned = "".join(ch for ch in _normalize_text(name) if ch not in _INVALID_SHEET_CHARS).strip()
    if not cleaned:
        raise ValueError("Название таблицы не может быть пустым.")
    return cleaned[:31]


def list_table_names(path: str | Path) -> list[str]:
    file_path = Path(path)
    if file_path.suffix.lower() != ".xlsx":
        raise ValueError("Файл итоговой таблицы должен быть в формате .xlsx.")
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _parse_table_from_sheet(sheet: Worksheet) -> DepreciationTable:
    table_title = _normalize_text(sheet.cell(row=1, column=4).value)
    contracts_label = _normalize_text(sheet.cell(row=2, column=2).value) or CONTRACTS_LABEL_DEFAULT

    period_to_column: dict[Period, int] = {}
    current_month_name: str | None = None
    for column in range(3, sheet.max_column + 1):
        row3_value = _normalize_text(sheet.cell(row=3, column=column).value)
        if row3_value:
            if row3_value == TOTAL_LABEL:
                current_month_name = None
                continue
            current_month_name = row3_value.lower()

        if current_month_name is None or current_month_name not in MONTH_TO_INDEX:
            continue

        year_value = sheet.cell(row=4, column=column).value
        if year_value in (None, ""):
            continue
        try:
            year = int(year_value)
        except (TypeError, ValueError):
            continue

        period_to_column[Period(year=year, month_index=MONTH_TO_INDEX[current_month_name])] = column

    contracts: list[ContractRow] = []
    row = 6
    while row <= sheet.max_row:
        label = _normalize_text(sheet.cell(row=row, column=2).value)
        if not label:
            row += 1
            continue
        if label.lower() == TOTAL_LABEL.lower():
            break

        values: dict[Period, object] = {}
        for period, column in period_to_column.items():
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value not in (None, ""):
                values[period] = cell_value
        contracts.append(ContractRow(contract_number=label, values=values))
        row += 1

    return DepreciationTable(
        table_title=table_title,
        contracts_label=contracts_label,
        periods=set(period_to_column.keys()),
        contracts=contracts,
    )


def load_depreciation_table(path: str | Path, table_name: str) -> DepreciationTable:
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        if table_name not in workbook.sheetnames:
            raise ValueError(f"В файле нет таблицы '{table_name}'.")
        return _parse_table_from_sheet(workbook[table_name])
    finally:
        workbook.close()


def create_empty_table(table_title: str) -> DepreciationTable:
    normalized = _normalize_text(table_title)
    if not normalized:
        raise ValueError("Название таблицы не может быть пустым.")
    return DepreciationTable(table_title=normalized)


def apply_contract_period_value(
    table: DepreciationTable,
    contract_number: str,
    period: Period,
    value: float | None,
) -> ApplyResult:
    period_created = period not in table.periods
    table.periods.add(period)

    row = table.find_contract(contract_number)
    contract_created = row is None
    if row is None:
        row = ContractRow(contract_number=_normalize_text(contract_number))
        table.contracts.append(row)

    value_written = False
    if value is not None and period not in row.values:
        row.values[period] = value
        value_written = True

    return ApplyResult(
        contract_number=contract_number,
        period=period,
        period_created=period_created,
        contract_created=contract_created,
        value_written=value_written,
    )


# Values meaning "no identifier available" - a literal dash, or the word "отсутствует"
# ("not available") that some contract registries use instead. Compared case-insensitively.
BLANK_IDENTIFIER_VALUES = {"-", "—", "–", "отсутствует"}

# Cyrillic letters that are visually identical to a Latin letter or digit - the same set
# used on Russian vehicle plates, plus З/3 which shows up interchangeably in VIN-like codes.
# Identifiers (VIN, serial, inventory number) get folded onto this common alphabet before
# comparison so 'ХЗW65392АН0001788' and 'X3W65392AH0001788' are recognized as the same value.
_CYRILLIC_LOOKALIKE_TRANSLATION = str.maketrans({
    "а": "a", "в": "b", "е": "e", "з": "3", "к": "k", "м": "m",
    "н": "h", "о": "o", "р": "p", "с": "c", "т": "t", "у": "y", "х": "x",
})


def _normalize_identifier(text: str) -> str:
    return text.casefold().translate(_CYRILLIC_LOOKALIKE_TRANSLATION)


@dataclass(frozen=True)
class AssetFailure:
    contract_number: str
    asset_label: str
    identifier_value: str
    reason: str


@dataclass(frozen=True)
class ContractMatchResult:
    contract_number: str
    total: float | None
    matched_count: int
    failures: list[AssetFailure]
    contract_error: str | None = None


def _load_generic_rows(path: str | Path) -> list[list[object]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("xlrd is required to read .xls files")
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(row_index) for row_index in range(sheet.nrows)]

    raise ValueError(f"Unsupported workbook format: {file_path.suffix}")


def _match_inventory_number_header(text: str) -> bool:
    normalized = re.sub(r"[\s.]+", "", text.lower())
    if not normalized.startswith("инв"):
        return False
    return "№" in normalized or "номер" in normalized or normalized.endswith("n")


def _match_vin_header(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text.lower())
    if "vin" in normalized:
        return True
    return "идентификацион" in normalized and "номер" in normalized


def _match_factory_number_header(text: str) -> bool:
    normalized = re.sub(r"[\s.]+", "", text.lower())
    if not normalized.startswith("заводск"):
        return False
    return "№" in normalized or "номер" in normalized


def _match_name_header(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text.lower())
    return normalized.startswith("наименован")


def _match_vedomost_inventory_header(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text.lower())
    return "инвентарн" in normalized and "номер" in normalized


def _match_vedomost_charge_header(text: str) -> bool:
    normalized = text.lower()
    return "начисл" in normalized and "амортиз" in normalized


@dataclass(frozen=True)
class _ContractIdentifierColumn:
    row: int
    column: int
    rule: str  # "inventory" or "search_all"


def _find_contract_identifier_column(rows: Sequence[Sequence[object]]) -> _ContractIdentifierColumn | None:
    inventory_hit: tuple[int, int] | None = None
    vin_hit: tuple[int, int] | None = None
    factory_hit: tuple[int, int] | None = None

    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            text = _normalize_text(value)
            if not text:
                continue
            if inventory_hit is None and _match_inventory_number_header(text):
                inventory_hit = (row_index, column_index)
            elif vin_hit is None and _match_vin_header(text):
                vin_hit = (row_index, column_index)
            elif factory_hit is None and _match_factory_number_header(text):
                factory_hit = (row_index, column_index)

    if inventory_hit is not None:
        return _ContractIdentifierColumn(row=inventory_hit[0], column=inventory_hit[1], rule="inventory")
    if vin_hit is not None:
        return _ContractIdentifierColumn(row=vin_hit[0], column=vin_hit[1], rule="search_all")
    if factory_hit is not None:
        return _ContractIdentifierColumn(row=factory_hit[0], column=factory_hit[1], rule="search_all")
    return None


def _find_vedomost_columns(rows: Sequence[Sequence[object]]) -> tuple[int, int, int | None] | None:
    """Returns (header_row, charge_column, inventory_column_or_None)."""
    for row_index, row in enumerate(rows):
        charge_column: int | None = None
        inventory_column: int | None = None
        for column_index, value in enumerate(row):
            text = _normalize_text(value)
            if not text:
                continue
            if _match_vedomost_charge_header(text):
                charge_column = column_index
            elif _match_vedomost_inventory_header(text):
                inventory_column = column_index
        if charge_column is not None:
            return row_index, charge_column, inventory_column
    return None


def _row_is_total_marker(row: Sequence[object]) -> bool:
    first_text = next((_normalize_text(value) for value in row if _normalize_text(value)), "")
    return first_text.casefold().startswith("итого")


def _match_by_inventory_number(identifier: str, statement_rows: Sequence[Sequence[object]], inventory_column: int, start_row: int) -> list[int]:
    target = _normalize_identifier(identifier)
    matches: list[int] = []
    for row_index in range(start_row, len(statement_rows)):
        row = statement_rows[row_index]
        if _row_is_total_marker(row):
            break
        if inventory_column >= len(row):
            continue
        cell_text = _normalize_identifier(_normalize_text(row[inventory_column]))
        if cell_text and cell_text == target:
            matches.append(row_index)
    return matches


def _match_by_substring(identifier: str, statement_rows: Sequence[Sequence[object]], start_row: int) -> list[int]:
    target = _normalize_identifier(identifier)
    matches: list[int] = []
    for row_index in range(start_row, len(statement_rows)):
        row = statement_rows[row_index]
        if _row_is_total_marker(row):
            break
        for value in row:
            cell_text = _normalize_identifier(_normalize_text(value))
            if cell_text and target in cell_text:
                matches.append(row_index)
                break
    return matches


def match_contract_depreciation(
    contract_number: str,
    contract_file: str | Path,
    statement_file: str | Path,
) -> ContractMatchResult:
    """Find the assets rented under `contract_file` inside `statement_file` and sum their
    "Начисление амортизации (износа)" / "За период" values.

    Matching rules, applied in order:
      1. If the contract registry has an "Инв. №"-like column, match its values against the
         "Инвентарный номер" column of the Ведомость (exact match).
      2. Otherwise, if it has an "Идентификационный номер (VIN)" or "Заводской №"-like
         column, search its values as a substring across every column of the Ведомость.
    Assets with a blank/dash identifier, no match, or an ambiguous (multi-row) match are
    reported as per-asset failures instead of silently affecting the total.
    """
    contract_file_name = Path(contract_file).name
    statement_file_name = Path(statement_file).name

    try:
        contract_rows = _load_generic_rows(contract_file)
    except Exception as exc:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=f"Не удалось прочитать файл договора '{contract_file_name}': {exc}",
        )

    id_info = _find_contract_identifier_column(contract_rows)
    if id_info is None:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=(
                f"В файле договора '{contract_file_name}' не найдена ни колонка 'Инв. №', "
                "ни колонка 'Идентификационный номер (VIN)' / 'Заводской №'. "
                "Обработка этого договора невозможна."
            ),
        )

    try:
        statement_rows = _load_generic_rows(statement_file)
    except Exception as exc:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=f"Не удалось прочитать файл Ведомости '{statement_file_name}': {exc}",
        )

    ved_info = _find_vedomost_columns(statement_rows)
    if ved_info is None:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=(
                f"В файле Ведомости '{statement_file_name}' не найдена колонка "
                "'Начисление амортизации (износа)'. Обработка этого договора невозможна."
            ),
        )
    ved_header_row, ved_charge_column, ved_inventory_column = ved_info

    if id_info.rule == "inventory" and ved_inventory_column is None:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=(
                f"В договоре '{contract_file_name}' используется колонка 'Инв. №', но в файле "
                f"Ведомости '{statement_file_name}' нет колонки 'Инвентарный номер'. "
                "Обработка этого договора невозможна."
            ),
        )

    header_row = id_info.row
    name_column: int | None = None
    for column_index, value in enumerate(contract_rows[header_row]):
        if _match_name_header(_normalize_text(value)):
            name_column = column_index
            break

    total = 0.0
    matched_count = 0
    failures: list[AssetFailure] = []

    for row_index in range(header_row + 1, len(contract_rows)):
        row = contract_rows[row_index]
        if _row_is_total_marker(row):
            break

        first_text = next((_normalize_text(value) for value in row if _normalize_text(value)), "")
        if not first_text:
            continue

        identifier_text = _normalize_text(row[id_info.column]) if id_info.column < len(row) else ""
        asset_name = _normalize_text(row[name_column]) if name_column is not None and name_column < len(row) else ""
        label = asset_name or f"строка {row_index + 1}"

        if not identifier_text or identifier_text.casefold() in BLANK_IDENTIFIER_VALUES:
            failures.append(AssetFailure(
                contract_number=contract_number,
                asset_label=label,
                identifier_value=identifier_text or "(пусто)",
                reason="значение идентификатора не указано (прочерк/пусто)",
            ))
            continue

        used_fallback_search = False
        if id_info.rule == "inventory":
            row_matches = _match_by_inventory_number(identifier_text, statement_rows, ved_inventory_column, ved_header_row + 1)
            if not row_matches:
                # Not found in the dedicated "Инвентарный номер" column - fall back to
                # searching the value across every column of the Ведомость.
                row_matches = _match_by_substring(identifier_text, statement_rows, ved_header_row + 1)
                used_fallback_search = True
        else:
            row_matches = _match_by_substring(identifier_text, statement_rows, ved_header_row + 1)

        distinct_rows = sorted(set(row_matches))

        if not distinct_rows:
            reason = "не найдено соответствие в Ведомости Амортизации"
            if used_fallback_search:
                reason += " (искали и в колонке 'Инвентарный номер', и по всем колонкам)"
            failures.append(AssetFailure(
                contract_number=contract_number,
                asset_label=label,
                identifier_value=identifier_text,
                reason=reason,
            ))
            continue

        if len(distinct_rows) > 1:
            reason = (
                "найдено несколько разных строк в Ведомости "
                f"(строки {', '.join(str(index + 1) for index in distinct_rows)}) - "
                "невозможно однозначно определить актив"
            )
            if used_fallback_search:
                reason += " (найдено при поиске по всем колонкам после того, как значение не нашлось в 'Инвентарный номер')"
            failures.append(AssetFailure(
                contract_number=contract_number,
                asset_label=label,
                identifier_value=identifier_text,
                reason=reason,
            ))
            continue

        matched_row = statement_rows[distinct_rows[0]]
        charge_value = matched_row[ved_charge_column] if ved_charge_column < len(matched_row) else None
        charge_number = _coerce_number(charge_value)
        if charge_number is None:
            charge_number = 0.0

        total += charge_number
        matched_count += 1

    if matched_count == 0 and not failures:
        return ContractMatchResult(
            contract_number=contract_number,
            total=None,
            matched_count=0,
            failures=[],
            contract_error=f"В реестре договора '{contract_file_name}' не найдено ни одной строки с активами.",
        )

    return ContractMatchResult(
        contract_number=contract_number,
        total=total if matched_count > 0 else None,
        matched_count=matched_count,
        failures=failures,
        contract_error=None,
    )


def _build_period_grid(periods: Iterable[Period]) -> list[Period]:
    months = sorted({period.month_index for period in periods})
    years = sorted({period.year for period in periods})
    return [Period(year=year, month_index=month_index) for month_index in months for year in years]


def _apply_block_border(sheet: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int, style: str = "thin") -> None:
    side = Side(style=style)
    border = Border(top=side, bottom=side, left=side, right=side)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            sheet.cell(row=row, column=column).border = border


def _apply_row_border(sheet: Worksheet, row: int, min_col: int, max_col: int, *, top: str | None = None, bottom: str | None = None) -> None:
    top_side = Side(style=top) if top else Side(style="thin")
    bottom_side = Side(style=bottom) if bottom else Side(style="thin")
    thin_side = Side(style="thin")
    for column in range(min_col, max_col + 1):
        sheet.cell(row=row, column=column).border = Border(top=top_side, bottom=bottom_side, left=thin_side, right=thin_side)


def _apply_row_font(sheet: Worksheet, row: int, min_col: int, max_col: int, *, bold: bool = False) -> None:
    for column in range(min_col, max_col + 1):
        sheet.cell(row=row, column=column).font = Font(bold=bold)


def _wide_merge_anchors(sheet: Worksheet) -> set[tuple[int, int]]:
    """(row, column) of every merge that spans more than one column, so its text isn't mistaken for that column's own content."""
    return {
        (merged_range.min_row, merged_range.min_col)
        for merged_range in sheet.merged_cells.ranges
        if merged_range.max_col > merged_range.min_col
    }


def _cell_text_width(cell: Cell) -> int:
    value = cell.value
    if value is None:
        return 0
    if isinstance(value, str) and value.startswith("="):
        return 0  # formula cells are sized via extra_widths instead, using the real computed value
    if isinstance(value, (int, float)) and cell.number_format == CURRENCY_FORMAT:
        return len(f"{value:,.2f} ₽")
    return len(str(value))


def _autosize_columns(
    sheet: Worksheet,
    *,
    min_column: int,
    max_column: int,
    extra_widths: dict[int, float] | None = None,
    min_width: int = 8,
    max_width: int = 60,
    padding: int = 2,
) -> None:
    extra_widths = extra_widths or {}
    skip = _wide_merge_anchors(sheet)
    for column in range(min_column, max_column + 1):
        max_len = 0
        for row in range(1, sheet.max_row + 1):
            if (row, column) in skip:
                continue
            max_len = max(max_len, _cell_text_width(sheet.cell(row=row, column=column)))
        extra = extra_widths.get(column)
        if extra is not None:
            max_len = max(max_len, len(f"{extra:,.2f} ₽"))
        width = min(max(max_len + padding, min_width), max_width)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_month_headers(sheet: Worksheet, period_grid: Sequence[Period]) -> dict[Period, int]:
    month_groups: list[tuple[str, list[Period]]] = []
    for period in period_grid:
        if month_groups and month_groups[-1][0] == period.month_name:
            month_groups[-1][1].append(period)
        else:
            month_groups.append((period.month_name, [period]))

    period_to_column: dict[Period, int] = {}
    column = 3
    for month_name, month_periods in month_groups:
        start_column = column
        for period in sorted(month_periods, key=lambda item: item.year):
            period_to_column[period] = column
            sheet.cell(row=4, column=column).value = period.year
            sheet.cell(row=5, column=column).value = AMORTIZATION_LABEL
            column += 1
        end_column = column - 1
        sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=end_column)
        sheet.cell(row=3, column=start_column).value = month_name

    return period_to_column


def _build_year_columns(period_grid: Sequence[Period], start_column: int) -> tuple[dict[int, int], int]:
    years = sorted({period.year for period in period_grid})
    year_to_column: dict[int, int] = {}
    column = start_column
    for year in years:
        year_to_column[year] = column
        column += 1
    return year_to_column, column - 1


def _write_total_headers(sheet: Worksheet, year_to_column: dict[int, int]) -> None:
    if not year_to_column:
        return
    start_column = min(year_to_column.values())
    end_column = max(year_to_column.values())
    sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=end_column)
    sheet.cell(row=3, column=start_column).value = TOTAL_LABEL
    for year, column in year_to_column.items():
        sheet.cell(row=4, column=column).value = year


def _render_into_sheet(sheet: Worksheet, table: DepreciationTable, tax_rules: Sequence[TaxRateRule]) -> None:
    if not table.contracts:
        raise ValueError("В таблице нет ни одного договора.")

    period_grid = _build_period_grid(table.periods)
    period_to_column = _write_month_headers(sheet, period_grid)
    last_month_column = max(period_to_column.values(), default=2)
    year_to_column, year_last_column = _build_year_columns(period_grid, last_month_column + 1)
    _write_total_headers(sheet, year_to_column)
    last_column = max(last_month_column, year_last_column)
    last_column_letter = get_column_letter(last_column)

    _apply_block_border(sheet, 1, 5, 1, last_column)

    sheet.merge_cells(f"D1:{last_column_letter}1")
    sheet["D1"] = table.table_title
    sheet["D1"].font = Font(bold=True)
    sheet["D1"].border = Border(top=Side(style="medium"), bottom=Side(style="thin"), left=Side(style="medium"), right=Side(style="medium"))

    sheet.merge_cells("A2:A5")
    sheet.merge_cells("B2:B5")
    sheet.merge_cells(f"C2:{last_column_letter}2")
    sheet["A2"] = COUNTERPARTY_LABEL
    sheet["B2"] = table.contracts_label
    sheet["C2"] = PROPERTY_LABEL

    column_extra: dict[int, float] = {}
    period_totals: dict[Period, float] = {period: 0.0 for period in period_to_column}
    year_totals: dict[int, float] = {year: 0.0 for year in year_to_column}

    row = 6
    contract_start = row
    for contract in table.contracts:
        sheet.cell(row=row, column=2).value = contract.contract_number

        for period, column in period_to_column.items():
            value = contract.values.get(period)
            if value is not None:
                cell = sheet.cell(row=row, column=column)
                cell.value = value
                cell.number_format = CURRENCY_FORMAT
                period_totals[period] += value

        for year, year_column in year_to_column.items():
            month_columns = [period_to_column[period] for period in period_grid if period.year == year]
            if not month_columns:
                continue
            terms = "+".join(f"{get_column_letter(column)}{row}" for column in month_columns)
            year_value = sum(contract.values.get(period) or 0 for period in period_grid if period.year == year)
            cell = sheet.cell(row=row, column=year_column)
            cell.value = f"={terms}"
            cell.number_format = CURRENCY_FORMAT
            column_extra[year_column] = max(column_extra.get(year_column, 0.0), abs(year_value))
            year_totals[year] += year_value

        _apply_row_border(sheet, row, 1, last_column)
        row += 1
    contract_end = row - 1

    total_row = row
    cleaned_row = row + 1
    sheet.cell(row=total_row, column=2).value = TOTAL_LABEL
    sheet.cell(row=cleaned_row, column=2).value = CLEANED_LABEL

    for period, column in list(period_to_column.items()):
        column_letter = get_column_letter(column)
        total_cell = sheet.cell(row=total_row, column=column)
        total_cell.value = f"=SUM({column_letter}{contract_start}:{column_letter}{contract_end})"
        total_cell.number_format = CURRENCY_FORMAT

        multiplier = tax_multiplier_for_year(period.year, tax_rules)
        cleaned_cell = sheet.cell(row=cleaned_row, column=column)
        cleaned_cell.value = f"={column_letter}{total_row}*{multiplier}"
        cleaned_cell.number_format = CURRENCY_FORMAT

        period_total = period_totals[period]
        column_extra[column] = max(column_extra.get(column, 0.0), abs(period_total), abs(period_total * multiplier))

    for year, column in year_to_column.items():
        column_letter = get_column_letter(column)
        total_cell = sheet.cell(row=total_row, column=column)
        total_cell.value = f"=SUM({column_letter}{contract_start}:{column_letter}{contract_end})"
        total_cell.number_format = CURRENCY_FORMAT

        multiplier = tax_multiplier_for_year(year, tax_rules)
        cleaned_cell = sheet.cell(row=cleaned_row, column=column)
        cleaned_cell.value = f"={column_letter}{total_row}*{multiplier}"
        cleaned_cell.number_format = CURRENCY_FORMAT

        year_total = year_totals[year]
        column_extra[column] = max(column_extra.get(column, 0.0), abs(year_total), abs(year_total * multiplier))

    _apply_row_font(sheet, total_row, 1, last_column, bold=True)
    _apply_row_font(sheet, cleaned_row, 1, last_column, bold=True)
    _apply_row_border(sheet, total_row, 1, last_column, top="medium")
    _apply_row_border(sheet, cleaned_row, 1, last_column, bottom="medium")

    _autosize_columns(sheet, min_column=1, max_column=last_column, extra_widths=column_extra)


def _replace_or_create_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    safe_name = _safe_sheet_title(sheet_name)
    if safe_name in workbook.sheetnames:
        index = workbook.sheetnames.index(safe_name)
        del workbook[safe_name]
        return workbook.create_sheet(title=safe_name, index=index)
    return workbook.create_sheet(title=safe_name)


def generate_depreciation_workbook(
    *,
    period: Period,
    statement_file: str | Path,
    contracts: Sequence[ContractInput],
    source_table_path: str | Path | None = None,
    source_table_name: str | None = None,
    new_table_name: str | None = None,
    tax_rules: Sequence[TaxRateRule] | None = None,
) -> GenerationResult:
    if not contracts:
        raise ValueError("Добавьте хотя бы один договор аренды.")

    if tax_rules is None:
        tax_rules = DEFAULT_TAX_RULES

    if source_table_path is not None:
        workbook = openpyxl.load_workbook(source_table_path, data_only=False)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

    messages: list[str] = []

    if source_table_name:
        if source_table_name not in workbook.sheetnames:
            raise ValueError(f"В файле нет таблицы '{source_table_name}'.")
        table = _parse_table_from_sheet(workbook[source_table_name])
        target_sheet_name = source_table_name
        messages.append(f"Дополняется существующая таблица '{source_table_name}'.")
    else:
        if not new_table_name or not new_table_name.strip():
            raise ValueError("Укажите название новой таблицы.")
        if source_table_path is not None and new_table_name.strip() in workbook.sheetnames:
            raise ValueError(f"В файле уже есть таблица '{new_table_name.strip()}'.")
        table = create_empty_table(new_table_name.strip())
        target_sheet_name = table.table_title
        messages.append(f"Создана новая таблица '{table.table_title}'.")

    period_already_existed = period in table.periods
    contract_messages: list[str] = []
    contract_errors: list[str] = []
    asset_failures: list[AssetFailure] = []

    for contract in contracts:
        match_result = match_contract_depreciation(contract.contract_number, contract.contract_file, statement_file)
        asset_failures.extend(match_result.failures)
        if match_result.contract_error:
            contract_errors.append(f"Договор '{contract.contract_number}' ({Path(contract.contract_file).name}): {match_result.contract_error}")

        apply_result = apply_contract_period_value(table, contract.contract_number, period, match_result.total)

        if apply_result.contract_created:
            row_note = "добавлена новая строка в таблице"
        elif apply_result.value_written:
            row_note = "записано новое значение"
        else:
            row_note = "строка уже существовала, значение не изменено"

        if match_result.total is not None:
            detail = f"учтено активов: {match_result.matched_count}"
            if match_result.failures:
                detail += f", не сопоставлено: {len(match_result.failures)}"
            contract_messages.append(f"Договор '{contract.contract_number}': {row_note} ({detail}).")
        elif match_result.contract_error:
            contract_messages.append(f"Договор '{contract.contract_number}': {row_note}, значение не рассчитано — см. список проблем.")
        else:
            contract_messages.append(f"Договор '{contract.contract_number}': {row_note}, ни один актив не сопоставлен — см. список проблем.")

    if period_already_existed:
        messages.append(f"Месяц {period.month_name} {period.year} уже существовал в таблице — новый столбец не добавлялся.")
    else:
        messages.append(f"Добавлен новый столбец периода: {period.month_name} {period.year}.")
    messages.extend(contract_messages)

    sheet = _replace_or_create_sheet(workbook, target_sheet_name)
    _render_into_sheet(sheet, table, tax_rules)
    workbook.active = workbook.sheetnames.index(sheet.title)

    return GenerationResult(
        workbook=workbook,
        sheet_name=sheet.title,
        messages=messages,
        contract_errors=contract_errors,
        asset_failures=asset_failures,
    )


def save_depreciation_workbook(result: GenerationResult, output_path: str | Path) -> None:
    result.workbook.save(output_path)
